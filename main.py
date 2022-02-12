from openpyxl import Workbook
import datetime
import sys
import openpyxl
import random
from openpyxl.styles import PatternFill

def processing(input_file, output_file, yc_row, ht_row):
    book = openpyxl.load_workbook(input_file)

    sheet = book.active

    my_red = openpyxl.styles.colors.Color(rgb='00FF0000')
    my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=my_red)

    row_index = 4
    yellow = "00FFFF00"
    while row_index <= sheet.max_row:
        yc_time = sheet[f'{yc_row}{row_index}'].value
        ht_time = sheet[f'{ht_row}{row_index}'].value
        # yc_time = datetime.datetime.strptime(yc_str, '%Y-%m-%d %H:%M:%S')
        # ht_time = datetime.datetime.strptime(ht_str, '%Y-%m-%d %H:%M:%S')
        diff = ht_time - yc_time
        second = diff.total_seconds()
        hour = second / 3600

        day = ht_time.date() - yc_time.date()
        hours = 2
        hours_added = datetime.timedelta(hours=hours)

        if yc_time.date() != ht_time.date() or yc_time > ht_time or hour < 2:
            hours = random.uniform(2, 5)
            hours_added = datetime.timedelta(hours=hours)
            ht_time = yc_time + hours_added
            sheet[f'{ht_row}{row_index}'] = ht_time
            sheet[f'{ht_row}{row_index}'].fill = PatternFill(start_color=yellow, end_color=yellow,
                                                      fill_type="solid")
            pass

        row_index += 1
        pass
    book.save(output_file)
    pass

if __name__ == '__main__':
    try:
        processing(sys.argv[1], sys.argv[2], sys.argv[3], sys.argv[4])
    except:
        print('main.exe input.xlsx output.xlsx M S')
        print('input.xlsx la ten file dau vao')
        print('output.xlsx la ten file dau ra')
        print('M la ten cot chua thoi diem yeu cau')
        print('S la ten cot chua thoi diem hoan thanh')

        pass

    pass




